"""What is on this campaign drive: logbooks, day folders, calibrations, space.

The first question asked of a drive that arrives in a box is not a scientific
one.  It is *which drive is this* -- how many observing days, which shot
numbers, whether the curator's logbook agrees, and whether there is room to
write cubes beside the data or the output has to go somewhere else.

Nothing here converts anything and nothing here writes to the drive.  One
record is measured and returned; the caller writes it wherever it belongs.

The curator's logbook is an ``.xlsx`` workbook, and reading one needs
``openpyxl``, which this package does not require.  A drive whose logbook
cannot be read is still worth inventorying, so a missing reader is reported as
a finding -- the logbook is present and unread -- and never as a failure.
"""

from __future__ import annotations

import json
import math
import re
import shutil
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

INVENTORY_SCHEMA = "echelle-drive-inventory/v1"

#: The curator names every workbook this way, one per experimental cycle.
LOGBOOK_GLOB = "*cycle*_Echelle_setup.xlsx"

#: ``000123_Echelle.SIF`` -- the shot number is the whole leading number.
SHOT_NAME = re.compile(r"^(\d+)_Echelle\.SIF$", re.IGNORECASE)

#: The trigger delay every ordinary day was taken with.  A day that used any
#: other one is called out by name, because its frames start elsewhere in the
#: discharge and nothing downstream can infer that from the files.
ORDINARY_TRIGGER_DELAY_S = 2.5

#: One exported cube is about this big, in gigabytes.  The estimate exists to
#: answer one question -- do the cubes fit beside the data -- so it is
#: deliberately a single round number rather than a model.
CUBE_GB_PER_SIF = 15 / 1000

#: Headroom demanded on top of the estimate before a drive is called roomy.
CUBE_HEADROOM_GB = 20.0

_DATE_COLUMN = "date"
_TRIGGER_COLUMN = "trig. delay(sec)"
_EXPOSURE_COLUMN = "exposure(sec)"
_FRAMES_COLUMN = "frames"

_NO_READER = (
    "openpyxl is not installed, so this logbook was not read "
    "(pip install openpyxl to read it)"
)


def _openpyxl() -> Any | None:
    """Return ``openpyxl`` if this installation has it, else ``None``.

    A named function rather than a bare import so the unread branch can be
    exercised on a machine that happens to have the reader.
    """

    try:
        import openpyxl
    except ImportError:
        return None
    return openpyxl


def _number(value: Any) -> float | None:
    """Read one cell as a number, or ``None`` when it is not one.

    Blank cells, text and the empty string are all *absent* rather than zero:
    a logbook row with no trigger delay recorded must not join the set of
    delays this drive was taken with.
    """

    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return None if math.isnan(number) else number
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _day_label(value: Any) -> str | None:
    """Read one ``date`` cell as the ``YYYYMMDD`` day the folders are named by."""

    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    number = _number(value)
    if number is None:
        return None
    return str(int(number))


def _workbook_rows(path: Path, openpyxl: Any) -> tuple[list[str], list[dict[str, Any]]]:
    """Return the first sheet's column names and its rows as dictionaries."""

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return [], []
        names = [str(cell).strip() if cell is not None else "" for cell in header]
        return names, [dict(zip(names, values)) for values in rows]
    finally:
        book.close()


def _odd_days(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Name every logbook day that was not taken at the ordinary trigger delay."""

    odd = []
    for row in rows:
        delay = _number(row.get(_TRIGGER_COLUMN))
        day = _day_label(row.get(_DATE_COLUMN))
        if delay is None or day is None or delay == ORDINARY_TRIGGER_DELAY_S:
            continue
        frames = _number(row.get(_FRAMES_COLUMN))
        odd.append(
            {
                "date": day,
                "trig_delay_s": delay,
                "exposure_s": _number(row.get(_EXPOSURE_COLUMN)),
                "frames": int(frames) if frames is not None else None,
            }
        )
    return odd


def read_logbook(path: Path) -> dict[str, Any]:
    """Read one curator logbook, or say in one field why it was not read."""

    openpyxl = _openpyxl()
    if openpyxl is None:
        return {"file": path.name, "unread": _NO_READER}
    try:
        names, rows = _workbook_rows(path, openpyxl)
        kept = (
            [row for row in rows if _day_label(row.get(_DATE_COLUMN)) is not None]
            if _DATE_COLUMN in names
            else rows
        )
        delays = sorted(
            {
                delay
                for delay in (_number(row.get(_TRIGGER_COLUMN)) for row in kept)
                if delay is not None
            }
        )
        days = [_day_label(row.get(_DATE_COLUMN)) for row in kept]
        date_range = [days[0], days[-1]] if _DATE_COLUMN in names and days else None
        return {
            "file": path.name,
            "rows": len(kept),
            "date_range": date_range,
            "trig_delays_seen": delays,
            "days_not_2p5s": _odd_days(kept) if _TRIGGER_COLUMN in names else [],
        }
    except Exception as exc:  # noqa: BLE001 - a logbook that will not parse is a finding
        # Deliberately wide: openpyxl raises half a dozen unrelated types over a
        # truncated, encrypted or simply mislabelled workbook, and none of them
        # is a reason to abandon the rest of the drive.
        return {"file": path.name, "error": str(exc)}


def _data_root(root: Path) -> Path:
    """Return the folder the day folders sit in: ``DATA`` when there is one."""

    data = root / "DATA"
    return data if data.is_dir() else root


def _is_calibration_folder(child: Path) -> bool:
    return "calib" in child.name.lower() or (child / "snapshot.toml").is_file()


def _day_record(child: Path) -> dict[str, Any]:
    """Count one day folder's frames and read its shot range off the names."""

    shots: list[int] = []
    sif_count = 0
    sample_size: int | None = None
    try:
        for entry in child.iterdir():
            if entry.name.startswith("."):
                # AppleDouble siblings (._000123_Echelle.SIF) a Mac writes on
                # exFAT are metadata, not frames.
                continue
            if entry.suffix.upper() == ".SIF":
                sif_count += 1
                if sample_size is None:
                    try:
                        sample_size = entry.stat().st_size
                    except OSError:
                        pass
            match = SHOT_NAME.match(entry.name)
            if match:
                shots.append(int(match.group(1)))
    except OSError as exc:
        return {"day": child.name, "error": str(exc)}
    return {
        "day": child.name,
        "sif_count": sif_count,
        "shot_range": [min(shots), max(shots)] if shots else None,
        "typical_file_bytes": sample_size,
    }


def _scan_days(base: Path, *, named_days_only: bool) -> tuple[list[dict], list[str]]:
    """Return one record per day folder and the calibration folder names.

    ``named_days_only`` is set when the scan fell back to the drive root
    because there is no ``DATA`` folder: a drive root holds cube folders,
    campaign pages and kit folders that are plainly not observing days, so
    there the name has to start with a four-digit date to count.  Under
    ``DATA`` every non-calibration folder is a day, which is what the drives
    in hand actually look like.
    """

    days: list[dict[str, Any]] = []
    calibrations: list[str] = []
    for child in sorted(base.iterdir()):
        if not child.is_dir() or child.name.startswith("."):
            continue
        if _is_calibration_folder(child):
            calibrations.append(child.name)
            continue
        if named_days_only and not child.name[:4].isdigit():
            continue
        days.append(_day_record(child))
    nested = base / "calibrations"
    if nested.is_dir():
        calibrations.extend(
            f"calibrations/{child.name}" for child in sorted(nested.iterdir()) if child.is_dir()
        )
    return days, calibrations


def inventory_drive(root: Path) -> dict[str, Any]:
    """Measure one mounted drive and return the record, writing nothing."""

    root = Path(root)
    usage = shutil.disk_usage(root)
    record: dict[str, Any] = {
        "schema": INVENTORY_SCHEMA,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "drive_root": str(root),
        "volume_name": root.name,
        "size_bytes": usage.total,
        "free_bytes": usage.free,
    }
    record["logbooks"] = [read_logbook(book) for book in sorted(root.glob(LOGBOOK_GLOB))]

    base = _data_root(root)
    record["data_root"] = str(base)
    if base.is_dir():
        days, calibrations = _scan_days(base, named_days_only=base == root)
    else:  # pragma: no cover - the caller resolves the root before calling
        days, calibrations = [], []
    record["days"] = days
    record["calibration_folders"] = calibrations
    return record


def cube_estimate_gb(record: dict[str, Any]) -> float:
    """Gigabytes of cubes this drive's frames would produce, roughly."""

    return sum(day.get("sif_count", 0) for day in record["days"]) * CUBE_GB_PER_SIF


def inventory_path(record: dict[str, Any], output_root: Path) -> Path:
    """Where this drive's record belongs under an output folder."""

    return Path(output_root) / "inventory" / f"{record['volume_name']}.json"


def write_inventory(record: dict[str, Any], output_root: Path) -> Path:
    """Write the record into ``<output_root>/inventory/`` and return its path."""

    path = inventory_path(record, output_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(record, indent=1), encoding="utf-8")
    return path


def _logbook_lines(logbooks: list[dict[str, Any]]) -> list[str]:
    """One line per curator logbook, plus a named line for each odd day."""

    if not logbooks:
        return ["logbook    : (none on this drive)"]
    lines = []
    for book in logbooks:
        if "error" in book:
            lines.append(f"logbook    : {book['file']} UNREADABLE: {book['error']}")
            continue
        if "unread" in book:
            lines.append(f"logbook    : {book['file']} PRESENT BUT UNREAD: {book['unread']}")
            continue
        span = "-".join(book["date_range"]) if book["date_range"] else "?"
        lines.append(
            f"logbook    : {book['file']}  {book['rows']} days  {span}  "
            f"delays {book['trig_delays_seen']}"
        )
        lines.extend(
            f"             ODD DAY {odd['date']}: trig delay {odd['trig_delay_s']} s, "
            f"exposure {odd['exposure_s']} s"
            for odd in book["days_not_2p5s"]
        )
    return lines


def format_inventory(record: dict[str, Any], written: Path | None = None) -> list[str]:
    """The summary an operator reads at the terminal, one line per fact."""

    free_gb = record["free_bytes"] / 1e9
    estimate = cube_estimate_gb(record)
    room = (
        "(cubes fit beside data)"
        if free_gb > estimate + CUBE_HEADROOM_GB
        else f"(TOO FULL for ~{estimate:.0f} GB of cubes -- output must go elsewhere)"
    )
    lines = [f"=== {record['volume_name']} ===", f"free space : {free_gb:.0f} GB {room}"]
    lines.extend(_logbook_lines(record["logbooks"]))
    days = record["days"]
    counted = [day for day in days if "error" not in day]
    frames = sum(day.get("sif_count", 0) for day in days)
    lines.append(
        f"day folders: {len(counted)}, SIF files: {frames}, est. cubes ~{estimate:.0f} GB"
    )
    if record["data_root"] == record["drive_root"]:
        # Say where the day folders were looked for, since it was not the
        # DATA folder the campaign drives normally carry.
        lines.append(f"note       : no DATA folder here; days were read from {record['data_root']}")
    first = next((day for day in days if day.get("shot_range")), None)
    last = next((day for day in reversed(days) if day.get("shot_range")), None)
    if first and last:
        lines.append(
            f"span       : {first['day']} (shot {first['shot_range'][0]}) .. "
            f"{last['day']} (shot {last['shot_range'][1]})"
        )
    folders = record["calibration_folders"]
    lines.append(f"calib dirs : {', '.join(folders) if folders else '(none on this drive)'}")
    for day in days:
        if "error" in day:
            lines.append(f"unreadable : {day['day']}: {day['error']}")
    if written is not None:
        lines.append(f"written    : {written}")
    return lines
